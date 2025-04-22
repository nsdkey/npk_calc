from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from .models import Fertilizer, P_profile
from django.db.models import Q
from .forms import FertilizerForm, P_profileForm, UploadFileForm
from scipy.optimize import minimize
import openpyxl
from django.http import HttpResponse, FileResponse
from django.conf import settings

# Create your views here.

def home(request):
    return render(request, 'main/home.html')

def download_p_profiles_example(request):
    return FileResponse((settings.MEDIA_ROOT / 'main' / 'p_profiles_example.xlsx').open('rb'), as_attachment=True)
def download_fertilizers_example(request):
    return FileResponse((settings.MEDIA_ROOT / 'main' / 'fertilizers_example.xlsx').open('rb'), as_attachment=True)

def user_login(request):
    er = {'error': None}
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('home')
        else:
            er['error'] = "Неверный логин или пароль"
    return render(request, 'main/login.html', er)

def user_logout(request):
    logout(request)
    return redirect('home')

def user_reg(request):
    er = {'error': None}
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        password2 = request.POST['password2']
        if password != '' and username != '':
            if password == password2:
                if username and not User.objects.filter(username=username).exists():
                    user = User.objects.create_user(username=username, password=password)
                    login(request, user)
                    return redirect('home')
                else:
                    er['error'] = "Пользователь с таким логином уже существует"
            else:
                er['error'] = "Пароли не совпадают"
        else:
            er['error'] = "Введите логин и пароль"
    return render(request, 'main/reg.html', er)

def profile(request):
    if request.user.is_authenticated:
        fertilizers = Fertilizer.objects.filter(Q(user = request.user))
        return render(request, 'main/profile.html', {'fertilizers': fertilizers})
    else:
        return redirect('home')

def profile_pp(request):
    if request.user.is_authenticated:
        p_profile = P_profile.objects.filter(Q(user = request.user))
        return render(request, 'main/profile_pp.html', {'p_profiles': p_profile})
    else:
        return redirect('home')
    
def delete_fertilizer(request, pk):
    get_object_or_404(Fertilizer, pk=pk).delete()
    return redirect('profile')

def add_fertilizer(request):
    if request.method == 'POST':
        form = FertilizerForm(request.POST)
        if form.is_valid():
            fertilizer = form.save(commit=False)
            fertilizer.user = request.user
            fertilizer.save()
            return redirect('profile')
    else:
        form = FertilizerForm()
    return render(request, 'main/add_fertilizer.html', {'form': form})

def edit_fertilizer(request, pk):
    fertilizer = get_object_or_404(Fertilizer, pk=pk)
    if request.method == 'POST':
        form = FertilizerForm(request.POST, instance=fertilizer)
        if form.is_valid():
            form.save()
            return redirect('profile')
    else:
        form = FertilizerForm(instance=fertilizer)
    return render(request, 'main/add_fertilizer.html', {'form': form})

def delete_p_profile(request, pk):
    get_object_or_404(P_profile, pk=pk).delete()
    return redirect('profile_pp')

def add_p_profile(request):
    if request.method == 'POST':
        form = P_profileForm(request.POST)
        if form.is_valid():
            p_profile = form.save(commit=False)
            p_profile.user = request.user
            p_profile.save()
            return redirect('profile_pp')
    else:
        form = P_profileForm()
    return render(request, 'main/add_p_profile.html', {'form': form})

def edit_p_profile(request, pk):
    p_profile = get_object_or_404(P_profile, pk=pk)
    if request.method == 'POST':
        form = P_profileForm(request.POST, instance=p_profile)
        if form.is_valid():
            form.save()
            return redirect('profile_pp')
    else:
        form = P_profileForm(instance=p_profile)
    return render(request, 'main/add_p_profile.html', {'form': form})

def calc(request):
    if request.user.is_authenticated:
        fertilizers = Fertilizer.objects.filter(Q(user=request.user) | Q(is_public=True))
        p_profiles = P_profile.objects.filter(Q(user=request.user) | Q(is_public=True))
    else:
        fertilizers = Fertilizer.objects.filter(Q(is_public=True))
        p_profiles = P_profile.objects.filter(Q(is_public=True))
    names = []
    answer = []
    er=""
    profile=[]
    el_names=[]
    selected_fertilizers = []
    selected_profile = p_profiles[0]
    if request.method == 'POST':
        selected_fertilizers = Fertilizer.objects.filter(id__in=request.POST.getlist('fertilizers'))
        selected_profile = get_object_or_404(P_profile, id=request.POST.get('p_profile'))
        water = float(request.POST.get('water_l'))

        if selected_fertilizers:
            f = []
            profile = list(vars(selected_profile).values())[6:]
            el_names = list(vars(selected_profile).keys())[6:]
            for i in range(8):
                profile[i+8] = profile[i+8] * 0.001

            for fertilize in selected_fertilizers:
                new_f = vars(fertilize)
                new_f['p2o5']=new_f['p2o5'] * 0.436
                new_f['k2o']=new_f['k2o'] * 0.83
                new_f['cao']=new_f['cao'] * 0.715
                new_f['mgo']=new_f['mgo'] * 0.603
                names.append(list(new_f.values())[4])
                f.append(list(new_f.values())[5:])

            def mix(doses, f):
                result = [0]*16
                for i in range(len(doses)):
                    for k in range(16):
                        result[k]+=(doses[i]*f[i][k])/100
                return result

            def func(doses, f, profile):
                fault = 0.05
                res = mix(doses, f)
                penalty = 0
                covered = 0
                for i in range(len(profile)):
                    if profile[i] > 0:
                        if abs(res[i]-profile[i])/profile[i] <= fault:
                            covered += 1
                        elif res[i] > profile[i]:
                            penalty += ((res[i] - profile[i]) ** 2) * (0.1+(((i//8))*0.9))
                        else:
                            penalty += (profile[i] - res[i]) * (0.1+(((i//8))*0.9))
                return penalty - (covered*(1/16)*penalty)
            
            start_doses = [0] * len(selected_fertilizers)
            bounds = [(0, None) for _ in range(len(selected_fertilizers))]

            answer = minimize(func, start_doses, args=(f, profile), bounds=bounds, method='L-BFGS-B').x 

            for k in range(16):
                for i in range(len(selected_fertilizers)):
                    profile[k] -= (f[i][k]*answer[i])/100
            profile = list(map(lambda x: round(x * water), profile))
            answer = list(map(lambda x: round(x * water), answer))

        else:
            er = "Выберите хотя бы 1 удобрение"

        
    return render(request, 'main/calc.html', {'fertilizers': fertilizers, 'p_profiles': p_profiles, 'answer': dict(zip(names, answer)), 'er': er, 'res': dict(zip(el_names, profile)), 'selected_fertilizers': [x.id for x in selected_fertilizers], 'selected_profile': selected_profile.id})

def calc_p_profile(request):
    start_profile = {
      "N": 100,
      "P": 20,
      "K": 50,
      "Ca": 60,
      "Mg": 25,
      "S": 15
    }
    return render(request, "main/calc_p_profile.html", {"macros": start_profile})

def upload_p_profiles(request):
    er = None
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        d = {'Старт': 'start', 'Вегетация': 'veg', 'Поздняя вегетация': 'late_veg', 'Плодоношение': 'fruiting'}
        if form.is_valid():
            file = request.FILES['file']
            tab = openpyxl.load_workbook(file).active
            for row in tab.iter_rows(min_row=2, values_only=True):
                if (type(row[0]) != str) or not(row[1] in d) or not(all(isinstance(x, (int, float)) for x in row[2:18])):
                    er = 'Некорректные данные в файле'
                    form = UploadFileForm()
                    return render(request, 'main/upload.html', {'form': form, 'error': er, 'cancel': 'profile_pp'})
            for row in tab.iter_rows(min_row=2, values_only=True):
                P_profile.objects.create(user=request.user, name=row[0],
                                        growth_stage=d.get(row[1]),
                                        no3=row[2], 
                                        nh4=row[3], 
                                        p=row[4], 
                                        k=row[5], 
                                        ca=row[6], 
                                        mg=row[7], 
                                        s=row[8], 
                                        cl=row[9], 
                                        fe=row[10], 
                                        mn=row[11], 
                                        b=row[12],
                                        zn=row[13], 
                                        cu=row[14], 
                                        mo=row[15], 
                                        co=row[16], 
                                        si=row[17])
            return redirect('profile_pp')
    else:
        form = UploadFileForm()
    
    return render(request, 'main/upload.html', {'form': form, 'error': er, 'cancel': 'profile_pp'})

def download_p_profiles(request):
    if request.user.is_authenticated:
        p_profiles = P_profile.objects.filter(Q(user = request.user))
        workbook = openpyxl.Workbook()
        tab = workbook.active
        tab.title = "Профили питания"
        tab.append(["Название", "Стадия роста", "NO3", "NH4", "P", "K", "Ca", "Mg", "S", "Cl", "Fe", "Mn", "B", "Zn", "Cu", "Mo", "Co", "Si"])
        for p_profile in p_profiles:
            tab.append([p_profile.name, p_profile.get_growth_stage_display(), p_profile.no3, p_profile.nh4, p_profile.p, p_profile.k, p_profile.ca, p_profile.mg, p_profile.s, p_profile.cl, p_profile.fe, p_profile.mn, p_profile.b, p_profile.zn, p_profile.cu, p_profile.mo, p_profile.co, p_profile.si])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename = "new_file.xlsx"'
        workbook.save(response)
        return response
    else:
        return redirect('home')

def upload_fertilizers(request):
    er = None
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            tab = openpyxl.load_workbook(file).active
            for row in tab.iter_rows(min_row=2, values_only=True):
                if (type(row[0]) != str) or not(all(isinstance(x, (int, float)) for x in row[1:17])):
                    er = 'Некорректные данные в файле'
                    form = UploadFileForm()
                    return render(request, 'main/upload.html', {'form': form, 'error': er, 'cancel': 'profile'})
            for row in tab.iter_rows(min_row=2, values_only=True):
                Fertilizer.objects.create(user=request.user, name=row[0],
                                        no3=row[1], 
                                        nh4=row[2], 
                                        p2o5=row[3], 
                                        k2o=row[4], 
                                        cao=row[5], 
                                        mgo=row[6], 
                                        s=row[7], 
                                        cl=row[8], 
                                        fe=row[9], 
                                        mn=row[10], 
                                        b=row[11],
                                        zn=row[12], 
                                        cu=row[13], 
                                        mo=row[14], 
                                        co=row[15], 
                                        si=row[16])
            return redirect('profile')
    else:
        form = UploadFileForm()
    
    return render(request, 'main/upload.html', {'form': form, 'error': er, 'cancel': 'profile'})

def download_fertilizers(request):
    if request.user.is_authenticated:
        p_profiles = Fertilizer.objects.filter(Q(user = request.user))
        workbook = openpyxl.Workbook()
        tab = workbook.active
        tab.title = "Удобрения"
        tab.append(["Название", "NO3", "NH4", "P2O5", "K2O", "CaO", "MgO", "S", "Cl", "Fe", "Mn", "B", "Zn", "Cu", "Mo", "Co", "Si"])
        for p_profile in p_profiles:
            tab.append([p_profile.name, p_profile.no3, p_profile.nh4, p_profile.p2o5, p_profile.k2o, p_profile.cao, p_profile.mgo, p_profile.s, p_profile.cl, p_profile.fe, p_profile.mn, p_profile.b, p_profile.zn, p_profile.cu, p_profile.mo, p_profile.co, p_profile.si])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename = "new_file.xlsx"'
        workbook.save(response)
        return response
    else:
        return redirect('home')